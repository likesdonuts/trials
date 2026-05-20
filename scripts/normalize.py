"""
normalize.py
------------
Reads all source court-data files from the workspace root and writes a
single normalized JSON file to app/public/data/trials.json.

Each source uses a different format:
  - Philadelphia Court of Common Pleas  → CSV
  - San Francisco Superior Court        → XLSX
  - Washoe County (Nevada)              → XLSX
  - SDNY Proceedings Calendar           → PDF (semi-structured text)
  - District of Massachusetts (Boston, Springfield, Worcester) → PDFs
  - District of Arizona                 → PDF

The output is a JSON array of Trial objects matching the TypeScript types
defined in src/types.ts.
"""

import csv
import json
import os
import re
import sys
import uuid
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl not found – install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import pdfplumber
except ImportError:
    print("pdfplumber not found – install with: pip install pdfplumber", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).parent.parent
OUT = ROOT / "app" / "public" / "data" / "trials.json"

# ---------------------------------------------------------------------------
# Court definitions
# ---------------------------------------------------------------------------

COURTS = {
    "phila": {
        "id": "phila-common-pleas",
        "name": "Philadelphia Court of Common Pleas",
        "shortName": "Phila. C.C.P.",
        "type": "state",
        "state": "PA",
    },
    "sf": {
        "id": "sf-superior",
        "name": "San Francisco Superior Court",
        "shortName": "S.F. Superior",
        "type": "state",
        "state": "CA",
    },
    "washoe": {
        "id": "washoe-district",
        "name": "Second Judicial District Court – Washoe County",
        "shortName": "Washoe Co. District",
        "type": "state",
        "state": "NV",
    },
    "sdny": {
        "id": "sdny",
        "name": "U.S. District Court – Southern District of New York",
        "shortName": "S.D.N.Y.",
        "type": "federal",
        "state": "NY",
    },
    "dma-boston": {
        "id": "dma-boston",
        "name": "U.S. District Court – District of Massachusetts (Boston)",
        "shortName": "D. Mass. (Boston)",
        "type": "federal",
        "state": "MA",
        "division": "Boston",
    },
    "dma-springfield": {
        "id": "dma-springfield",
        "name": "U.S. District Court – District of Massachusetts (Springfield)",
        "shortName": "D. Mass. (Springfield)",
        "type": "federal",
        "state": "MA",
        "division": "Springfield",
    },
    "dma-worcester": {
        "id": "dma-worcester",
        "name": "U.S. District Court – District of Massachusetts (Worcester)",
        "shortName": "D. Mass. (Worcester)",
        "type": "federal",
        "state": "MA",
        "division": "Worcester",
    },
    "daz": {
        "id": "daz",
        "name": "U.S. District Court – District of Arizona",
        "shortName": "D. Ariz.",
        "type": "federal",
        "state": "AZ",
    },
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_id(source_key: str, index: int) -> str:
    return f"{source_key}-{index:04d}"


def to_iso(dt) -> str:
    """Return YYYY-MM-DD for a date, datetime, or parseable string."""
    if isinstance(dt, (datetime,)):
        return dt.date().isoformat()
    if isinstance(dt, date):
        return dt.isoformat()
    if isinstance(dt, str):
        for fmt in ("%d-%b-%Y", "%m/%d/%Y", "%Y-%m-%d", "%B %d, %Y"):
            try:
                return datetime.strptime(dt.strip(), fmt).date().isoformat()
            except ValueError:
                pass
    return str(dt)


def classify_status(result_type: str | None) -> str:
    if not result_type:
        return "scheduled"
    rt = result_type.lower()
    if "settl" in rt:
        return "settled"
    if "vacat" in rt:
        return "vacated"
    return "scheduled"


TRIAL_KEYWORDS = re.compile(
    r"\bjury\s+trial\b|\bbench\s+trial\b|\bnon.jury\s+trial\b|"
    r"\btrial\s+-\s+jury\b|\btrial\s+-\s+non.?jury\b|"
    r"\btrial\s+-\s+short\s+jury\b|\btrial\s+day\b",
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# 1. Philadelphia Court of Common Pleas  (CSV)
# ---------------------------------------------------------------------------

def parse_philadelphia(path: Path) -> list[dict]:
    records = []
    court = COURTS["phila"]
    with open(path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for i, row in enumerate(reader):
            trial_date_raw = row.get("Trial Date Certain", "").strip()
            if not trial_date_raw:
                continue
            records.append({
                "id": make_id("phila", i),
                "docketNumber": row.get("Case ID", "").strip(),
                "caseName": row.get("Caption", "").strip().title(),
                "trialDate": to_iso(trial_date_raw),
                "court": court,
                "eventType": row.get("Court Type", "Trial").strip().title(),
                "status": "scheduled",
                "sourceFile": path.name,
            })
    return records


# ---------------------------------------------------------------------------
# 2. San Francisco Superior Court  (XLSX)
# ---------------------------------------------------------------------------

def parse_san_francisco(path: Path) -> list[dict]:
    records = []
    court = COURTS["sf"]
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if not row or not row[0]:
            continue
        rec = dict(zip(headers, row))
        raw_date = rec.get("Date")
        if raw_date is None:
            continue
        # Fix obvious year typo: 2027 should be 2026
        iso = to_iso(raw_date)
        if iso.startswith("2027-"):
            iso = "2026-" + iso[5:]
        time_val = rec.get("Time")
        trial_time = None
        if hasattr(time_val, "hour"):
            trial_time = f"{time_val.hour:02d}:{time_val.minute:02d}"
        dept = rec.get("Dept")
        records.append({
            "id": make_id("sf", i),
            "docketNumber": str(rec.get("Case Number", "")).strip(),
            "caseName": str(rec.get("Case Title", "")).strip().title(),
            "trialDate": iso,
            **({"trialTime": trial_time} if trial_time else {}),
            "court": court,
            "eventType": "Trial",
            "department": str(dept) if dept else None,
            "status": "scheduled",
            "sourceFile": path.name,
        })
    return records


# ---------------------------------------------------------------------------
# 3. Washoe County, Nevada  (XLSX)
# ---------------------------------------------------------------------------

def parse_washoe(path: Path) -> list[dict]:
    records = []
    court = COURTS["washoe"]
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if not row or not row[0]:
            continue
        rec = dict(zip(headers, row))
        event_type = str(rec.get("Event Type", "")).strip()
        if not TRIAL_KEYWORDS.search(event_type):
            continue
        raw_dt = rec.get("Event Date/Time")
        if raw_dt is None:
            continue
        iso = to_iso(raw_dt)
        trial_time = None
        if isinstance(raw_dt, datetime):
            trial_time = f"{raw_dt.hour:02d}:{raw_dt.minute:02d}"
        # Extract judge name from Role/Name column
        role_name = str(rec.get("Role/Name", ""))
        judge = None
        m = re.search(r"District Judge:\s+([^,]+),\s+(\S+)", role_name)
        if m:
            judge = f"{m.group(2).title()} {m.group(1).title()}"
        dept = rec.get("Department")
        status = classify_status(rec.get("Result Type"))
        # Strip trailing department label like "(D1)" or "(D3)" from case names
        raw_case = str(rec.get("Case Name", "")).strip()
        raw_case = re.sub(r"\s*\([Dd]\d+\)\s*$", "", raw_case)
        raw_case = re.sub(r"\s*\([Dd]\d+$", "", raw_case)  # unclosed paren
        records.append({
            "id": make_id("washoe", i),
            "docketNumber": str(rec.get("Case Number", "")).strip(),
            "caseName": raw_case.strip().title(),
            "trialDate": iso,
            **({"trialTime": trial_time} if trial_time else {}),
            "court": court,
            "eventType": event_type,
            **({"judge": judge} if judge else {}),
            **({"department": str(dept)} if dept else {}),
            "status": status,
            "sourceFile": path.name,
        })
    return records


# ---------------------------------------------------------------------------
# 4. SDNY Proceedings Calendar  (PDF)
#
# Line format (approximately):
#   {Case Title} {Docket#} {EventType ...} {Judge} {Location} {MM/DD/YYYY} {HH:MM AM/PM}
#
# We only keep lines whose EventType contains a trial keyword.
# ---------------------------------------------------------------------------

SDNY_DOCKET = re.compile(
    r"\b(\d{2}-(?:CV|CR|MC|MD|MI|MP|MG|MJ|SC|EP|SP)-\d+)\b", re.IGNORECASE
)
SDNY_DATE = re.compile(r"\b(\d{2}/\d{2}/\d{4})\b")
SDNY_TIME = re.compile(r"\b(\d{1,2}:\d{2}\s*(?:AM|PM))\b", re.IGNORECASE)

SDNY_LOCATIONS = {
    "DPM": "Daniel Patrick Moynihan Courthouse",
    "TM": "Thurgood Marshall Courthouse",
    "WP": "White Plains Courthouse",
}


def parse_sdny(path: Path) -> list[dict]:
    records = []
    court = COURTS["sdny"]
    with pdfplumber.open(path) as pdf:
        full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)

    # Each entry spans one logical line; PDF may wrap, so we join continuation
    # lines (lines that don't start with a date/time pattern).
    lines = full_text.splitlines()
    merged: list[str] = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Skip footer/legend lines
        if re.match(r"^(DPM|TM|WP)\s*=", line):
            continue
        if re.match(r"^(Week of:|Last Updated|Page \d|Civil & Criminal)", line):
            continue
        # If the line starts with a time token it's a new entry (D.Mass style)
        # or a date token (SDNY style); otherwise it may be a continuation
        if merged and not re.match(r"^\d{1,2}:\d{2}", line) and not SDNY_DATE.search(line):
            # Check if previous line already has a date – if so, this is new
            if not SDNY_DATE.search(merged[-1]):
                merged[-1] += " " + line
                continue
        merged.append(line)

    idx = 0
    current_date = None
    for line in merged:
        # Try to pick up a date from the line
        dm = SDNY_DATE.search(line)
        if dm:
            current_date = to_iso(dm.group(1))
        if not TRIAL_KEYWORDS.search(line):
            continue
        if not current_date:
            continue
        docket_m = SDNY_DOCKET.search(line)
        docket = docket_m.group(1) if docket_m else ""
        time_m = SDNY_TIME.search(line)
        trial_time = None
        if time_m:
            t = datetime.strptime(time_m.group(1).strip(), "%I:%M %p")
            trial_time = f"{t.hour:02d}:{t.minute:02d}"
        # Case name = everything before the docket number
        case_name = line
        if docket_m:
            case_name = line[: docket_m.start()].strip()
        case_name = re.sub(r"\s+", " ", case_name).strip().title()
        # Event type = text after docket up to judge name (heuristic)
        event_type = "Trial"
        if docket_m:
            after_docket = line[docket_m.end():].strip()
            et_m = re.match(r"((?:Jury|Bench|Non-Jury)\s+Trial(?:\s+Day\s+\d+)?)", after_docket, re.IGNORECASE)
            if et_m:
                event_type = et_m.group(1).strip().title()
        # Judge = word(s) between event type and location abbrev / date
        judge = None
        judge_m = re.search(
            r"(?:Jury|Bench|Non-Jury)\s+Trial(?:\s+Day\s+\d+)?\s+(\S+(?:\s+\S+)??)(?:\s+(?:DPM|TM|WP)\b|\s+\d{2}/)",
            line,
        )
        if judge_m:
            judge = judge_m.group(1).strip()
        records.append({
            "id": make_id("sdny", idx),
            "docketNumber": docket,
            "caseName": case_name,
            "trialDate": current_date,
            **({"trialTime": trial_time} if trial_time else {}),
            "court": court,
            "eventType": event_type,
            **({"judge": judge} if judge else {}),
            "status": "scheduled",
            "sourceFile": path.name,
        })
        idx += 1
    return records


# ---------------------------------------------------------------------------
# 5. District of Massachusetts  (PDFs)
#
# Line format (approximately, organized by judge section):
#   {HH:MM AM/PM}  {Docket#}  {Party A}  v.  {Party B}  {EventType}
#   Judge header:  "Judge {Name} - Courtroom {X} - {N}th Floor"
# ---------------------------------------------------------------------------

DMASS_DOCKET = re.compile(
    r"\b(\d:\d{2}-(?:CV|CR|MD)-\d{4,6}(?:-\d+)?)\b", re.IGNORECASE
)
DMASS_TIME = re.compile(r"^(\d{1,2}:\d{2}\s*(?:AM|PM))\s+", re.IGNORECASE)
DMASS_JUDGE = re.compile(
    r"^(?:Judge|Senior Judge|Chief Judge|Magistrate Judge|Chief Magistrate Judge)\s+(.+?)\s+-\s+",
    re.IGNORECASE,
)
DMASS_DATE_HDR = re.compile(r"Calendar\s+-\s+\w+,\s+(\w+\s+\d+,\s+\d{4})")


def _parse_dmass_single(path: Path, court_key: str) -> list[dict]:
    records = []
    court = COURTS[court_key]
    with pdfplumber.open(path) as pdf:
        full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)

    current_date = None
    current_judge = None
    idx = 0
    lines = full_text.splitlines()

    # Merge continuation lines (lines that don't start with a time or judge keyword)
    merged: list[str] = []
    for line in lines:
        line = line.rstrip()
        if not line.strip():
            continue
        if re.match(r"^Page \d+ of \d+", line.strip()):
            continue
        if (DMASS_TIME.match(line.strip()) or DMASS_JUDGE.match(line.strip())
                or DMASS_DATE_HDR.search(line) or re.match(r"^(Last Updated|United States District Court|District of Massachusetts|Calendar)", line.strip())):
            merged.append(line.strip())
        elif merged:
            merged[-1] += " " + line.strip()
        else:
            merged.append(line.strip())

    for line in merged:
        # Detect calendar date from header
        date_m = DMASS_DATE_HDR.search(line)
        if date_m:
            current_date = to_iso(date_m.group(1))
            continue
        # Detect judge
        judge_m = DMASS_JUDGE.match(line)
        if judge_m:
            current_judge = judge_m.group(1).strip()
            continue
        # Detect event line
        time_m = DMASS_TIME.match(line)
        if not time_m:
            continue
        if not TRIAL_KEYWORDS.search(line):
            continue
        if not current_date:
            continue
        t_str = time_m.group(1).strip()
        trial_time = None
        try:
            t = datetime.strptime(t_str, "%I:%M %p")
            trial_time = f"{t.hour:02d}:{t.minute:02d}"
        except ValueError:
            pass
        rest = line[time_m.end():].strip()
        docket_m = DMASS_DOCKET.search(rest)
        docket = docket_m.group(1) if docket_m else ""
        # Case name: between docket and event type keyword
        case_name = rest
        if docket_m:
            case_name = rest[docket_m.end():].strip()
        # Strip event type and everything after it
        et_m = re.search(
            r"((?:Jury|Bench|Non-Jury)\s+Trial(?:\s+Day\s+\d+)?)",
            case_name,
            re.IGNORECASE,
        )
        event_type = "Trial"
        if et_m:
            event_type = et_m.group(1).strip().title()
            case_name = case_name[: et_m.start()].strip()
        case_name = re.sub(r"\s+", " ", case_name).strip().title()
        # Courtroom from judge line (captured in current judge context)
        dept = None
        if current_judge:
            dept_m = re.search(r"Courtroom\s+(\S+)", line)
            if not dept_m:
                dept_m = re.search(r"Courtroom\s+(\S+)", f"Judge {current_judge} - Courtroom ?")
        records.append({
            "id": make_id(court_key.replace("-", "_"), idx),
            "docketNumber": docket,
            "caseName": case_name,
            "trialDate": current_date,
            **({"trialTime": trial_time} if trial_time else {}),
            "court": court,
            "eventType": event_type,
            **({"judge": current_judge} if current_judge else {}),
            "status": "scheduled",
            "sourceFile": path.name,
        })
        idx += 1
    return records


def parse_dmass_boston(path: Path) -> list[dict]:
    return _parse_dmass_single(path, "dma-boston")


def parse_dmass_springfield(path: Path) -> list[dict]:
    return _parse_dmass_single(path, "dma-springfield")


def parse_dmass_worcester(path: Path) -> list[dict]:
    return _parse_dmass_single(path, "dma-worcester")


# ---------------------------------------------------------------------------
# 6. District of Arizona  (PDF)
#
# Format: organized by judge/courtroom sections; each event line:
#   {HH:MM AM/PM}  {Party Name(s)}  {Docket#}  {EventType}
# ---------------------------------------------------------------------------

DAZ_DOCKET = re.compile(
    r"\b((?:CR|CV|MJ|MC)\s+\d{2}-\d{4,6}(?:-\d{2}-[A-Z]{3}-[A-Z]{3}(?:-[A-Z]{3})?)?)\b",
    re.IGNORECASE,
)
DAZ_DATE_HDR = re.compile(r"(\d{1,2}/\d{1,2}/\d{2,4}),?\s+\d{1,2}:\d{2}\s+[AP]M")
DAZ_JUDGE = re.compile(r"^Honorable\s+(.+?)\s+-\s+", re.IGNORECASE)
DAZ_TIME = re.compile(r"^(\d{1,2}:\d{2}\s*(?:AM|PM))\s+", re.IGNORECASE)


def parse_daz(path: Path) -> list[dict]:
    records = []
    court = COURTS["daz"]
    with pdfplumber.open(path) as pdf:
        full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)

    # The PDF header contains the date in the format "5/20/26, 3:30 PM"
    doc_date = None
    date_hdr_m = re.search(r"(\d{1,2}/\d{1,2}/\d{2}),\s+\d", full_text)
    if date_hdr_m:
        raw = date_hdr_m.group(1)
        parts = raw.split("/")
        if len(parts) == 3:
            month, day, year = parts
            if len(year) == 2:
                year = "20" + year
            doc_date = f"{int(year):04d}-{int(month):02d}-{int(day):02d}"

    current_judge = None
    idx = 0
    lines = full_text.splitlines()
    merged: list[str] = []
    for line in lines:
        line = line.rstrip()
        if not line.strip():
            continue
        if re.match(r"https?://", line.strip()):
            continue
        if (DAZ_TIME.match(line.strip()) or DAZ_JUDGE.match(line.strip())):
            merged.append(line.strip())
        elif merged:
            merged[-1] += " " + line.strip()
        else:
            merged.append(line.strip())

    for line in merged:
        judge_m = DAZ_JUDGE.match(line)
        if judge_m:
            current_judge = judge_m.group(1).strip()
            current_judge = re.sub(r"\s+-\s+.*$", "", current_judge)
            continue
        time_m = DAZ_TIME.match(line)
        if not time_m:
            continue
        if not TRIAL_KEYWORDS.search(line):
            continue
        if not doc_date:
            continue
        t_str = time_m.group(1).strip()
        trial_time = None
        try:
            t = datetime.strptime(t_str, "%I:%M %p")
            trial_time = f"{t.hour:02d}:{t.minute:02d}"
        except ValueError:
            pass
        rest = line[time_m.end():].strip()
        docket_m = DAZ_DOCKET.search(rest)
        docket = docket_m.group(1).strip() if docket_m else ""
        case_name = rest
        if docket_m:
            case_name = rest[: docket_m.start()].strip()
        # Trim trailing event type from case name
        et_m = re.search(
            r"((?:Jury|Bench)\s+Trial(?:\s+Day\s+\d+)?)",
            case_name,
            re.IGNORECASE,
        )
        event_type = "Jury Trial"
        if et_m:
            event_type = et_m.group(1).strip().title()
            case_name = case_name[: et_m.start()].strip()
        else:
            et_m2 = re.search(r"((?:Jury|Bench)\s+Trial(?:\s+Day\s+\d+)?)", rest, re.IGNORECASE)
            if et_m2:
                event_type = et_m2.group(1).strip().title()
        # Strip trailing stray tokens (single words, ampersands, etc.)
        case_name = re.sub(r"\s+", " ", case_name).strip()
        case_name = re.sub(r"\s+&\s*$", "", case_name).strip()
        case_name = case_name.title()
        records.append({
            "id": make_id("daz", idx),
            "docketNumber": docket,
            "caseName": case_name,
            "trialDate": doc_date,
            **({"trialTime": trial_time} if trial_time else {}),
            "court": court,
            "eventType": event_type,
            **({"judge": current_judge} if current_judge else {}),
            "status": "scheduled",
            "sourceFile": path.name,
        })
        idx += 1
    return records


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    all_trials: list[dict] = []

    parsers = [
        (ROOT / "Philadelphia Court of Common Pleas - Civil Court Listings Results.csv", parse_philadelphia),
        (ROOT / "San Francisco Superior Court.xlsx", parse_san_francisco),
        (ROOT / "Washoe County Nevada 5_20_26.xlsx", parse_washoe),
        (ROOT / "SDNY Proceedings Calendar 5.18.2026.pdf", parse_sdny),
        (ROOT / "D Mass Boston 5_20_26.pdf", parse_dmass_boston),
        (ROOT / "D Mass Boston 5_21_26.pdf", parse_dmass_boston),
        (ROOT / "D Mass Springfield 5_21_26.pdf", parse_dmass_springfield),
        (ROOT / "D Mass Worcester 5_21_26.pdf", parse_dmass_worcester),
        (ROOT / "D Arizona 5_20_26.pdf", parse_daz),
    ]

    for path, parser in parsers:
        if not path.exists():
            print(f"WARNING: {path} not found, skipping", file=sys.stderr)
            continue
        try:
            results = parser(path)
            print(f"  {path.name}: {len(results)} trial(s)")
            all_trials.extend(results)
        except Exception as exc:
            print(f"ERROR parsing {path.name}: {exc}", file=sys.stderr)
            import traceback; traceback.print_exc()

    # Sort by date then by court name
    all_trials.sort(key=lambda t: (t["trialDate"], t["court"]["name"]))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(all_trials, fh, indent=2, default=str)

    print(f"\nWrote {len(all_trials)} total trials → {OUT}")


if __name__ == "__main__":
    main()
